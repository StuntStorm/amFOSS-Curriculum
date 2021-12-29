import sys
import openpyxl

def blankRowInserter(index, blanks, filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    rows = tuple(sheet.rows)

    for rowObj in rows[::-1]:
        for cellObj in rowObj:
            c = cellObj.column
            r = cellObj.row
            if r >= index and r < index+blanks:
                sheet.cell(row=r+blanks, column=c).value = cellObj.value
                sheet.cell(row=r, column=c).value = ''
            elif r >= index+blanks:
                sheet.cell(row=r+blanks, column=c).value = cellObj.value

    wb.save('after_'+filename)
num_args = len(sys.argv)

if num_args < 4:
    print("usage: python3 blankRowInserter.py 3 2 NameOfTheFile.xlsx")
else:
    index = int(sys.argv[1])
    blanks = int(sys.argv[2])
    filename = sys.argv[3]
    blankRowInserter(index, blanks, filename)
