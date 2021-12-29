import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter

if len(sys.argv) > 1:
	n = int(sys.argv[1])
else:
	n = 6

wb = openpyxl.Workbook()
sheet = wb.active

boldFont = Font(bold=True)
for i in range(2, n+2):
	multiply = i - 1
	column = 'A' + str(i)
	sheet[column] = multiply
	sheet[column].font = boldFont
	row = get_column_letter(i) + '1'
	sheet[row] = multiply
	sheet[row].font = boldFont

for i in range(2, n+2):
	leftvalue = i - 1
	for s in range(2, n+2):
		topvalue = s - 1
		cell = get_column_letter(s) + str(i)
		sheet[cell] = leftvalue*topvalue

sheet.freeze_panes = 'B2'

wb.save('multiplicationTable.xlsx')
