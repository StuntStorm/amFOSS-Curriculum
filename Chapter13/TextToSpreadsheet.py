import os
import openpyxl

def textToSheet(directory, filename):
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='result')
    sheet = wb.active

    colIndex = 1
    for file in os.listdir():
        if file.endswith('.txt'):
            rowIndex = 1
            with open(file) as r:
                for line in r:
                    sheet.cell(row=rowIndex, column=colIndex).value = line
                    rowIndex += 1
            colIndex += 1
    wb.save(filename)


direct = input("Enter the Directory : ")
savename = input("Enter name to save : ")
textToSheet(direct, savename+".xlsx")
