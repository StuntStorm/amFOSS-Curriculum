import openpyxl

def invertCells(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    newSheet = wb.create_sheet(index=0, title='inverted')
    for rowvalue in sheet.rows:
        for cellvalue in rowvalue:
            colIndex = cellvalue.column
            rowIndex = cellvalue.row
            newSheet.cell(row=colIndex, column=rowIndex).value = cellvalue.value
    wb.save('invert_'+filename)

invert = input("Enter the File name : ")
invertCells(invert+".xlsx")
